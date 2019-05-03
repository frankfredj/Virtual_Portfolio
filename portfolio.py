import openpyxl
import os.path
import datetime

from pandas_datareader import data
import pandas as pd
from openpyxl.utils import get_column_letter
import numpy as np

import urllib
from bs4 import BeautifulSoup as bs
import requests

import bs4
import re
import logging
import more_itertools

from selenium.webdriver import Chrome
from selenium import webdriver 
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
import win32clipboard
from openpyxl.styles import Alignment
import time

import matplotlib.pyplot as plt
from sklearn.decomposition import PCA

from tqdm import tqdm
from tqdm import tnrange

from scipy.interpolate import interp1d

from scipy.stats import norm

from dateutil.parser import parse

import seaborn as sns


class Portfolio:


	def __init__(self, directory, file_name):


		if file_name[-4] + file_name[-3] + file_name[-2] + file_name[-1] != "xlsx":
			file_name = file_name + ".xlsx"

		file_location = os.path.join(directory, file_name)


		if not os.path.isfile(file_location):

			#Constructor in case the xlsx file needs to be created

			wb = openpyxl.Workbook()

			wb["Sheet"].title = "Prices"
			wb.create_sheet("Stocks")
			wb.create_sheet("Dates")
			wb.create_sheet("Accounts")
			wb.create_sheet("Rates")

			date_columns = ["Date", "Hour", "Min", "Sec", "Assets", "Debt", "Net assets"]
			for j in range(7):

				wb["Dates"][get_column_letter(j+1) + str(1)].value = date_columns[j]


			wb["Accounts"]["A1"].value = 0.0
			wb["Accounts"]["B1"].value = 0.0
			wb["Rates"]["A1"].value = 0.05
			wb["Rates"]["B1"].value = 0.025

			wb.save(file_location)

			self.stocks = np.array([])
			self.shares = np.array([])
			self.equity = np.array([])

			self.borrowing_rate = 0.05
			self.sick_free_rate = 0.025

			self.assets = pd.DataFrame(np.zeros(shape = (1,3)))
			self.assets.columns = ["Equity", "Debt", "Net assets"]

			self.directory = directory
			self.file_name = file_name
			self.file_location = file_location

			self.xlsx = wb

			self.file_location_2 = os.path.join(directory, file_name + "_Analysis")

			self.xlsx_2 = wb_analysis



			print("Empty portfolio created.")


		else:

			#Constructor based on an already existing xlsx file

			wb = openpyxl.load_workbook(file_location)

			s = wb["Stocks"]

			self.stocks = []
			self.shares = np.zeros(wb["Stocks"].max_row-1)

			j = 0
			

			for  i in range(wb["Stocks"].max_row-1):

				self.stocks.append(wb["Stocks"]["A" + str(i+2)].value) 
				self.shares[j] = wb["Stocks"]["B" + str(i+2)].value
				j += 1

			self.stocks = np.array(self.stocks)

			n = wb["Prices"].max_row

			self.equity = np.zeros(len(self.stocks))



			for j in range(len(self.stocks)):

				try:
					self.equity[j] = wb["Prices"][get_column_letter(j+1) + str(n)].value * self.shares[j]

				except:
					print(self.stocks[j] + " couldn't be priced...")


			cash = np.zeros(shape = (1,3))
			cash[0,0] = wb["Accounts"]["A1"].value
			cash[0,1] = wb["Accounts"]["B1"].value
			cash[0,2] = cash[0,0] + cash[0,1]

			self.assets = pd.DataFrame(cash)
			self.assets.columns = ["Equity", "Debt", "Net assets"]

			self.directory = directory
			self.file_name = file_name
			self.file_location = file_location

			self.xlsx = wb

			self.borrowing_rate = self.xlsx["Rates"]["A1"].value
			self.risk_free_rate = self.xlsx["Rates"]["B1"].value

			print("Existing portfolio loaded.")

			self.update()


	#Mutators for the rates

	def set_borrowing_rate(self, r):

		self.borrowing_rate = r
		self.xlsx["Rates"]["A1"].value = r

	def set_risk_free_rate(self,r):

		self.risk_free_rate = r
		self.xlsx["Rates"]["B1"].value = r

	def set_risk_free_rate_t_bill(self):

		try:

			self.get_T_bill_rates()

			n = np.shape(self.t_bill_rates)[0] - 1

			self.risk_free_rate = self.t_bill_rates.iloc[n,5] / 100
			self.xlsx["Rates"]["B1"].value = self.risk_free_rate

			print("")

			print("Risk free rate updated to " + str(self.t_bill_rates.iloc[n,5]) + "%")

		except:

			print("Something went wrong: couldn't get the T_bill rate")


	#-------------------------------

	#Excell utility functions

	#-------------------------------


	#Attempts to convert a string to a datetime object

	#returns 0 or 1 if it fails 
	#0: invalid string format
	#1: invalid input type

	def parse_if_needed(self, obj):

		if type(obj) == str:

			try :

				obj = parse(obj)

			except:

				#print("Couldn't recognize string format.")
				return 0

			else:

				return obj

		else:

			if type(obj) == datetime.datetime or type(obj) == pd._libs.tslibs.timestamps.Timestamp:

				return obj

			elif type(obj) == np.datetime64:

				return pd.Timestamp(obj)


			else:

				#print("Input is neither a string nor a datetime object.")
				return 1



	#Returns the date difference in seconds

	#Returns 0,1 or 2 in case of an error
	#0: date_1 is wrong
	#1: date_2 is wrong
	#2: both are wrong 

	def date_diff_seconds(self, date_1, date_2):

		date_1 = self.parse_if_needed(date_1)
		date_2 = self.parse_if_needed(date_2)

		cond_1 = type(date_1) != int
		cond_2 = type(date_2) != int

		if cond_1 and cond_2:

			return (date_1.date() - date_2.date()).total_seconds()

		else:

			if not cond_1 and not cond_2:				

				print("")
				print("ERROR: date_1 and date_2 both have problematic formatings.")
				return 2

			elif not cond_1:

				print("")
				print("ERROR: date_1 has a problematic formating.")
				return 0

			else:

				print("")
				print("ERROR: date_2 has a problematic formating.")
				return 1


	#Finds the first non-empty sheet amongst a set with sheetnames "self.stocks[i]" + "subset"


	def find_first_non_empty_sheet(self, subset):

		k = 1

		for i in self.stocks:

			sheet = i + subset

			if self.xlsx[sheet].max_row >= 2:

				break

			k += 1

		if k == len(self.stocks):

			print("All sheets are empty.")
			return 0

		else:

			return sheet


	#Finds the date range amongst a subset of sheets
	#Used to update Yahoo closing data

	def find_date_range(self, subset):

		
		sheet = self.find_first_non_empty_sheet(subset)

		m = self.xlsx[sheet].max_column
		n = self.xlsx[sheet].max_row

		k = 1

		for i in range(m):

			if not type(self.parse_if_needed(self.xlsx[sheet][get_column_letter(i+1) + str(2)].value)) == int:

				break

			k += 1

		if not type(self.parse_if_needed(self.xlsx[sheet][get_column_letter(k) + str(2)].value)) == int:

			return [self.parse_if_needed(self.xlsx[sheet][get_column_letter(k) + str(2)].value).date()  , self.parse_if_needed(self.xlsx[sheet][get_column_letter(k) + str(n)].value).date()]

		else:

			print("ERROR: no dates founds")

			return 0








	#Quick search amongst an ordered date list with no duplicated elements
	#RETURNS THE CLOSEST MATCH IN TERMS OF SHEET INDEX (integer)

	#This function is useful to merge sheets together.
	#The search algorithm shouldn't make too many evaluations
	#(n_evals is bounded by 4 + error made by first guess)


	def find_date_amongst_ordered(self, date, sheet, column):

		if type(column) != str:

			column = get_column_letter(int(column))

		#Checks if the provided date lies within the sheet's date range
		#Handles the problem by returning either the lowest or highest date (aka closest match)

		n = self.xlsx[sheet].max_row

		date = self.parse_if_needed(date)

		date_min = self.parse_if_needed(self.xlsx[sheet][column + str(2)].value)
		date_max = self.parse_if_needed(self.xlsx[sheet][column + str(n)].value)

		delta_1 = self.date_diff_seconds(date, date_min)
		delta_2 = self.date_diff_seconds(date_max, date)



		if delta_1 <= 0:

			#print("DATE NOT WITHIN RANGE: returning first date instead.")
			return 2

		elif delta_2 <= 0:

			#print("DATE NOT WITHIN RANGE: returning last date instead.")
			return n



		#Searching algorithm

		else:

			#Formulate a first guess based on the distances between our date and the minimum / maximum sheet dates

			guess = 1 + (n-1) * delta_1 / (delta_1 + delta_2)
			guess = int(guess)

			#Handle cases where the guess is the first or last sheet date

			if guess <= 2:

				return 2

			elif guess >= n:

				return n

			else:

				#Initialise variables needed to search above and below the initial guess

				up = guess
				down = guess

				date_up = self.xlsx[sheet][column + str(up)].value
				date_down = self.xlsx[sheet][column + str(down)].value

				#Declare lists that will store the time differences as we search alongside both directions
				#Start with the initial guess

				up_vals = [np.abs(self.date_diff_seconds(date, date_up))]				
				down_vals = [up_vals[0]]

				#While the distance between our date and our next guess diminishes (i.e.: we are getting closer), keep searching
				#Do this in both directions

				while up < n :

					up += 1
					date_up = self.xlsx[sheet][column + str(up)].value
					up_vals.append(np.abs(self.date_diff_seconds(date, date_up)))

					if up_vals[-1] > up_vals[-2]:

						up -= 1
						break


				while down > 2:

					down -= 1
					date_down = self.xlsx[sheet][column + str(down)].value
					down_vals.append(np.abs(self.date_diff_seconds(date, date_down)))

					if down_vals[-1] > down_vals[-2]:

						down += 1
						break

				#Check for which direction provided the best guess based on absolute time difference in seconds
				#Return the proper index

				if np.min(up_vals) < np.min(down_vals):

					return up

				else:

					return down



	def find_group_of_dates(self, date, sheet, column):

		try:

			if type(column) != str:

				column = get_column_letter(int(column))

			guess = self.find_date_amongst_ordered(date, sheet, column)

			n = self.xlsx[sheet].max_row

			up = guess
			down = guess

			date_up = [self.xlsx[sheet][column + str(up)].value]
			date_down = [date_up[0]]


			while up < n :

				up += 1
				date_up.append(self.xlsx[sheet][column + str(up)].value)

				if date_up[-1] != date_up[-2]:

					up -= 1
					break


			while down > 2:

				down -= 1
				date_down.append(self.xlsx[sheet][column + str(down)].value) 

				if date_down[-1] != date_down[-2]:

					down += 1
					break

		except:

			print("ERROR: something went wrong.")
			return -1

		else:

			return np.arange(down, up + 1)



	def extract_sheet_by_date_group(self, date, sheet, column):

		try:

			rows = self.find_group_of_dates(date, sheet, column)
			columns = np.arange(1, self.xlsx[sheet].max_column + 1)

			rows2 = []
			for i in rows:
				rows2.append(int(i))

			columns2 = []
			for i in columns:
				columns2.append(int(i))	

			frame = self.extract_xlsx_sheet_over_range(sheet, columns2, rows2)

		except:

			print("ERROR: something went wrong.")
			return -1

		else:

			return frame


	#Appends a pandas DataFrame to an excel sheet
	#Column dimentions have to match
	#Includes column names if sheet is empty

	def append_frame_to_sheet(self, sheet, data_frame):


		if sheet in self.xlsx.sheetnames and (np.shape(data_frame)[1] == self.xlsx[sheet].max_column or self.xlsx[sheet].max_column == 1):

			n = self.xlsx[sheet].max_row + 1

			n_f = np.shape(data_frame)[0]
			m = np.shape(data_frame)[1]

			for i in range(n_f):
				for j in range(m):

					self.xlsx[sheet][get_column_letter(j+1) + str(n + i)].value = data_frame.iloc[i,j]	

			#print("")
			#print("Successfully appended new data.")

			if n == 2:

				for j in range(m):

					self.xlsx[sheet][get_column_letter(j+1) + str(1)].value = data_frame.columns.values[j]


			self.xlsx.save(self.file_location)

		else:

			if not sheet in self.xlsx.sheetnames:

				print("ERROR: Couldn't find any sheet named " + sheet + ".")

			elif not np.shape(data_frame)[1] == self.xlsx[sheet].max_column:

				dim_col = np.shape(data_frame)[1]
				sheet_col = self.xlsx[sheet].max_column

				print("ERROR: Sheet has " + str(sheet_col) + " columns; provided frame has " + str(dim_col) + ".")



	#Merges a pandas Data Frame to an excel sheet based on dates

	#c_instance: if there are N columns with dates, then which one should we base our indexing upon?
	#if_same: if the dates are perfect matches, what should we do? "replace", "append", "nothing"

	def incorporate_frame_to_sheet_by_dates(self, sheet, c_instance , data_frame, if_same):


		if sheet in self.xlsx.sheetnames and np.shape(data_frame)[1] == self.xlsx[sheet].max_column:

			m = self.xlsx[sheet].max_column

			i = 0
			index = -1
			instance = 0

			while i < m:

				i += 1

				if not type(self.parse_if_needed(self.xlsx[sheet][get_column_letter(i) + str(2)].value)) == int:

					index = i
					instance += 1

					if instance == c_instance:

						break

						#print("")
						#print("Date index located...")
						#print("")



			if index == -1:

				print("ERROR: no columns containing dates were found.")

			elif type(self.parse_if_needed(data_frame.iloc[0, index -1])) == int:

				print("ERROR: date indices do not match")

			else:

				#print("Transfering data...")
				#print("")

				for j in tqdm(range(np.shape(data_frame)[0])):

					try:

						date = data_frame.iloc[j,index - 1]

						closest_date_on_sheet_index = self.find_date_amongst_ordered(date, sheet, index)
						closest_date_on_sheet = self.parse_if_needed(self.xlsx[sheet][get_column_letter(index) + str(closest_date_on_sheet_index)].value)

						diff = (self.parse_if_needed(date).date() - closest_date_on_sheet.date()).total_seconds()

					except:

						continue

					else:

						if diff < 0: 

							self.xlsx[sheet].insert_rows(closest_date_on_sheet_index, amount = 1)

							for k in range(np.shape(data_frame)[1]):
								self.xlsx[sheet][get_column_letter(k+1) + str(closest_date_on_sheet_index)].value = data_frame.iloc[k,index - 1]

						elif diff > 0:

							self.xlsx[sheet].insert_rows(closest_date_on_sheet_index + 1, amount = 1)

							for k in range(np.shape(data_frame)[1]):
								self.xlsx[sheet][get_column_letter(k+1) + str(closest_date_on_sheet_index + 1)].value = data_frame.iloc[k,index - 1]

						elif diff == 0:

							if if_same == "replace":

								for k in range(np.shape(data_frame)[1]):
									self.xlsx[sheet][get_column_letter(k+1) + str(closest_date_on_sheet_index)].value = data_frame.iloc[k,index - 1]

							elif if_same == "append":

								self.xlsx[sheet].insert_rows(closest_date_on_sheet_index, amount = 1)

								for k in range(np.shape(data_frame)[1]):
									self.xlsx[sheet][get_column_letter(k+1) + str(closest_date_on_sheet_index)].value = data_frame.iloc[k,index - 1]

				self.xlsx.save(self.file_location)


	#Extract frame from an excel sheet

	#return 0 or 1 if it fails
	#0: bad formating on xlsx part
	#1: sheet is empty


	def extract_xlsx_sheet(self, sheet):

		n = self.xlsx[sheet].max_row
		m = self.xlsx[sheet].max_column

		if n > 1:

			frame = []

			for i in range(n-1):

				new_row = []

				for j in range(m):

					new_row.append(self.xlsx[sheet][get_column_letter(j+1) + str(2+i)].value)

				frame.append(new_row)	

			try: 

				columns = []

				for j in range(m):

					columns.append(self.xlsx[sheet][get_column_letter(j+1) + str(1)].value)


				frame = pd.DataFrame(data = frame, columns = columns)


			except:

				return 0

			else:


				return frame

		else:

			return 1




	def extract_xlsx_sheet_over_range(self, sheet, cols, rows):

		n = self.xlsx[sheet].max_row
		m = self.xlsx[sheet].max_column

		cols = [x for x in cols if isinstance(x, int)]
		rows = [x for x in rows if isinstance(x, int)]

		cols = [x for x in cols if 0 < x < m+1]
		rows = [x for x in rows if 1 < x < n+1]

		if n > 1 and len(cols) > 0 and len(rows) > 0:

			frame = []

			for i in rows:

				new_row = []

				for j in cols:

					new_row.append(self.xlsx[sheet][get_column_letter(j) + str(i)].value)

				frame.append(new_row)	

			try: 

				columns = []

				for j in cols:

					columns.append(self.xlsx[sheet][get_column_letter(j) + str(1)].value)


				frame = pd.DataFrame(data = frame, columns = columns)


			except:

				return 0

			else:


				return frame

		else:

			return 1


	#Dependency
	#Used to match extracted dates indices

	def match_date_index(self, dates):

		indices = []

		for k in range(len(dates)):

			dates[k] = np.array(dates[k])
			indices.append(np.arange(0, len(dates[k])))

		for k in range(len(dates) - 1):

			index = np.intersect1d(dates[k], dates[k+1], return_indices = True)[1]

			for kk in range(k+2):

				indices[kk] = indices[kk][index,]
				dates[kk] = dates[kk][index,]

		return indices


	#Extracts a column from each data frame
	#Data is grouped according to matching dates
	#Rows with missing data are dropped

	#Typical use: properly retrieve adj close price frame

	def extract_grouped_data(self, subset, data_index, date_index):

		frame = []
		dates = []
		columns = []

		for i in tqdm(self.stocks):

			sheet = i + subset

			rows = []

			for j in range(self.xlsx[sheet].max_row+1):

				rows.append(j)


			col = self.extract_xlsx_sheet_over_range(sheet, [data_index], rows)
			date = self.extract_xlsx_sheet_over_range(sheet, [date_index], rows)

			if not isinstance(col, int) and not isinstance(date, int):

				frame.append(col)
				dates.append(np.array(date))
				columns.append(i)

		if len(frame) > 0:

			dates_0 = dates[0]

			try:

				index = self.match_date_index(dates)

				for i in range(len(frame)):

					frame[i] = frame[i].iloc[index[i],:]

				for i in range(len(frame)-1):

					frame[0][columns[i+1]] = frame[i+1]

				frame = frame[0]

				frame.rename(columns={frame.columns[0] : columns[0]}, inplace=True)

				rows = []

				for k in index[0]:

					rows.append(self.parse_if_needed(dates_0[k][0]).date())

				frame.index = rows

				frame.dropna(inplace=True)

			except:

				print("ERROR: something went wrong.")	

				return 0	

			else:

				return frame	

		else:

				print("ERROR: something went wrong.")	

				return 0



	#Extracts grouped data with matching date indices over a range of date

	def extract_grouped_data_over_range(self, subset, data_index, date_index, frm, to):

		frame = []
		dates = []
		columns = []

		for i in tqdm(self.stocks):

			sheet = i + subset

			start = self.find_date_amongst_ordered(frm, sheet, 7)
			end = self.find_date_amongst_ordered(to, sheet, 7)

			rows = []

			if end - start > 0:

				for j in range(end - start + 1):

					rows.append(start + j)

			else:

				continue


			col = self.extract_xlsx_sheet_over_range(sheet, [data_index], rows)
			date = self.extract_xlsx_sheet_over_range(sheet, [date_index], rows)

			if not isinstance(col, int) and not isinstance(date, int):

				frame.append(col)
				dates.append(np.array(date))
				columns.append(i)

		if len(frame) > 0:

			dates_0 = dates[0]

			try:

				index = self.match_date_index(dates)

				for i in range(len(frame)):

					frame[i] = frame[i].iloc[index[i],:]

				for i in range(len(frame)-1):

					frame[0][columns[i+1]] = frame[i+1]

				frame = frame[0]

				frame.rename(columns={frame.columns[0] : columns[0]}, inplace=True)

				rows = []

				for k in index[0]:

					rows.append(self.parse_if_needed(dates_0[k][0]).date())

				frame.index = rows

				frame.dropna(inplace=True)

			except:

				print("ERROR: something went wrong.")	

				return 0	

			else:

				return frame	

		else:

				print("ERROR: something went wrong.")	

				return 0





	#-------------------------------

	#Excell utility functions

	#-------------------------------




	#-------------------------------

	#Data Scrapping utility functions

	#-------------------------------


	def get_T_bill_rates(self):

		#Scrape date from www.treasury.gov

		url ="https://www.treasury.gov/resource-center/data-chart-center/interest-rates/Pages/TextView.aspx?data=yield"
		page = requests.get(url)
		soup = bs(page.text, "html.parser")

		frame = pd.read_html(str(soup.find_all("table", {"class" : "t-chart"})))[0]

		if "T_rates" in self.xlsx.sheetnames:
			self.xlsx.remove(self.xlsx["T_rates"])

		self.xlsx.create_sheet("T_rates")

		self.append_frame_to_sheet("T_rates", frame)

		self.xlsx.save(self.file_location)
		self.t_bill_rates = frame

		print("T-Bill rates were succesfully updated.")




	def update_option_price(self, ticker):

		if ticker in self.stocks:

			self.xlsx.remove(self.xlsx[ticker + "Calls"])
			self.xlsx.remove(self.xlsx[ticker + "Puts"])

			self.xlsx.create_sheet(ticker + "Calls")
			self.xlsx.create_sheet(ticker + "Puts")
		
			#get the unix time stamp of the possible expiration dates
			#uses a headless chrome driver to locate the drop down menu

			url = "https://finance.yahoo.com/quote/" + ticker + "/options"

			print("")
			print("Extracting option exercise dates for " + ticker + "...")

			options = Options()
			options.add_argument("--headless")

			driver = webdriver.Chrome(options = options)

			try:

				driver.get(url)

			except:
				drive.close()
				print("ERROR: couldn't access webpage for stock " + ticker)

			else:

				expiration_dates = driver.find_element_by_xpath("//select").text
				expiration_dates = expiration_dates.split("\n")

				exp_datetime = []

				#Yahoo uses UNIX timestamps based on the seconds elapsed from Jan 1 1970

				for i in expiration_dates:
					exp_datetime.append(datetime.datetime.strptime(i, "%B %d, %Y").date())

				unix_stamps = []

				for i in exp_datetime:
					stamp = i - datetime.date(1970,1,1)
					unix_stamps.append(int(stamp.total_seconds()))

				driver.close()


				#get data for each UNIX stamps

				print("")
				print("Extracting data tables for " + ticker + "...")

				index = 0

				for stamps in tqdm(unix_stamps):

					url_2 = url + "?date=" + str(stamps)

					try:

						page = requests.get(url_2)
						soup = bs(page.text, "html.parser")

						tables = soup.find_all("table", {"class" : "calls"})
						call = pd.read_html(str(tables))[0].iloc[:,1:]
						call["Expiration"] = expiration_dates[index]


						tables = soup.find_all("table", {"class" : "puts"})
						put = pd.read_html(str(tables))[0].iloc[:,1:]
						put["Expiration"] = expiration_dates[index]

						self.append_frame_to_sheet(ticker + "Calls", call)
						self.append_frame_to_sheet(ticker + "Puts", put)

						index += 1				

					except:

						
						print("ERROR: " + ticker + " options expiring on " + expiration_dates[index] + " could not be loaded.")

						index += 1

				print("")
				print(ticker + " option data scrapping completed.")

				self.xlsx.save(self.file_location)

		else:
			print(ticker + " not found within portfolio")



	def update_option_price_list(self,tickers):

		for i in tickers:

			try: 
				self.update_option_price(i)

			except:
				continue



	def fill_missing_historical_data(self):

		date_range = self.find_date_range("Historical_data")

		frm = date_range[0].strftime("%Y-%m-%d")
		to = date_range[1].strftime("%Y-%m-%d")

		for ticker in tqdm(self.stocks):

			try:

				if self.xlsx[ticker + "Historical_data"].max_row == 1:

					frame = data.DataReader(ticker, "yahoo", frm, to)
					frame["Dates"] = frame.index

					self.append_frame_to_sheet(ticker + "Historical_data", frame)

				else:
					continue

			except:
				continue

		self.xlsx.save(self.file_location)





	def replace_all_historical_data(self, start_date, end_date):


		start_date = self.parse_if_needed(start_date)
		end_date = self.parse_if_needed(end_date)


		if not isinstance(start_date, str) and not isinstance(end_date, str):

			start_date = start_date.strftime("%Y-%m-%d")
			end_date = end_date.strftime("%Y-%m-%d")

			for i in tqdm(self.stocks):

				try:

					frame = data.DataReader(i, 'yahoo', start_date, end_date)
					frame["Dates"] = frame.index

					self.xlsx.remove(self.xlsx[i + "Historical_data"])
					self.xlsx.create_sheet(i + "Historical_data")

					self.append_frame_to_sheet(i + "Historical_data", frame)

				except:

					print(i + " couldn't be loaded...")

			self.xlsx.save(self.file_location)

			print("")
			print("Stock prices were successfully replaced.")

		else:

			print("")
			print("ERROR: couldn't decypher provided dates")

			return 0




	def bridge_historical_data(self):

		now = datetime.datetime.now()
		now_str = now.strftime("%Y-%m-%d")

		for i in tqdm(self.stocks):

			n = self.xlsx[i + "Historical_data"].max_row

			try:

				last_entry =  self.xlsx[i + "Historical_data"]["G" + str(n)].value

				if isinstance(last_entry, pd._libs.tslibs.timestamps.Timestamp):
					last_entry_string = last_entry.strftime("%Y-%m-%d")


				else:
					last_entry_string = self.parse_if_needed(last_entry).strftime("%Y-%m-%d")
				 

			except:

				continue


			if last_entry_string != now_str:

				try:

					new_data = data.DataReader(i, 'yahoo', last_entry_string, now_str)
					new_data["Dates"] = new_data.index

				except:

					continue

				else:

					k = np.shape(new_data)[0]

					while new_data.index[0].strftime("%Y-%m-%d") == last_entry_string and k > 1:

						new_data = new_data.iloc[1:,:]
						k -= 1

					if k == 1 and new_data.index[0].strftime("%Y-%m-%d") == last_entry_string:

						continue				

					else:

						index = np.array(new_data.index.strftime("%Y-%m-%d"))
						index = np.unique(index, return_index = True)[1]

						new_data = new_data.iloc[index,:]

						try:

							self.append_frame_to_sheet(i + "Historical_data", new_data)

						except:

							continue


		self.xlsx.save(self.file_location)

		print("")
		print("Historical prices have been bridged up to today's date.")





	def scrape_option_greeks_1(self, url):

		try:

			page = requests.get(url)
			soup = bs(page.text, "html.parser")

			data = pd.read_html(str(soup.find_all("table")))[2]

			split_index = np.where(data.columns.values == "Puts")[0][0]

			calls = data.iloc[:,:split_index]
			puts = data.iloc[:,split_index:]

			puts["Strike"] = calls["Strike"]

			return {"Calls" : calls, "Puts" : puts}


		except:

			print("ERROR: couldn't access Yahoo webpage for stock " + ticker + ".")

			return 0



	def scrape_option_greeks(self, ticker, nmonths):

		url = "https://www.nasdaq.com/symbol/" + ticker + "/option-chain/greeks"

		url = url + "?dateindex=" + str(nmonths)

		count = 0
		calls = []
		puts = []

		stop = False

		while not stop:

			count += 1
			url_2 = url + "&page=" + str(count)

			frame = self.scrape_option_greeks_1(url_2)

			if type(frame) == int:

				break

			if count > 1:

				last = calls[count - 2]["Calls"][0]
				current = frame["Calls"]["Calls"][0]

				if last == current:

					break

				else:

					calls.append(frame["Calls"])
					puts.append(frame["Puts"])


			else:

				calls.append(frame["Calls"])
				puts.append(frame["Puts"])


		if len(calls) > 0:

			calls = pd.concat(calls)
			puts = pd.concat(puts)

			return {"Calls" : calls, "Puts" : puts}

		else:

			print("ERROR: couldn't access Yahoo webpage for stock " + ticker + ".")

			return 0



	def update_greeks(self, ticker, nmonths):


		if nmonths >= 0 and ticker in self.stocks:

			nmonths = int(nmonths)

			self.xlsx.remove(self.xlsx[ticker + "Greeks_Calls"])
			self.xlsx.remove(self.xlsx[ticker + "Greeks_Puts"])

			self.xlsx.create_sheet(ticker + "Greeks_Calls")
			self.xlsx.create_sheet(ticker + "Greeks_Puts")

			for i in tqdm(range(nmonths+1)):

				try:

					data = self.scrape_option_greeks(ticker, i)

					data["Calls"]["Dates"] = datetime.datetime.now().date()
					data["Puts"]["Dates"] = datetime.datetime.now().date()					

					self.append_frame_to_sheet(ticker + "Greeks_Calls", data["Calls"])
					self.append_frame_to_sheet(ticker + "Greeks_Puts", data["Puts"])

				except:

					continue

			print("")
			print("Scrapping completed for " + ticker + " option Greeks.")

			self.xlsx.save(self.file_location)


		else:

			print("ERROR: nonsensical argument provided for \"nmonths\"")


	#-------------------------------

	#Data Scrapping utility functions

	#-------------------------------




	#-------------------------------

	#.xslx Data extractors

	#-------------------------------


	def extract_historical_data(self, ticker):

		return self.extract_xlsx_sheet(ticker + "Historical_data")

	def extract_calls(self, ticker):

		return self.extract_xlsx_sheet(ticker + "Calls")

	def extract_puts(self, ticker):

		return self.extract_xlsx_sheet(ticker + "Calls")

	def extract_greeks_calls(self, ticker):

		return self.extract_xlsx_sheet(ticker + "Greeks_Calls")

	def extract_greeks_puts(self, ticker):

		return self.extract_xlsx_sheet(ticker + "Greeks_Puts")



	def extract_adjusted_close(self):

		return self.extract_grouped_data("Historical_data", 6, 7)


	def extract_adjusted_close_over_range(self, frm, to):

		return self.extract_grouped_data_over_range("Historical_data", 6, 7, frm, to)


	def extract_greeks_calls_date(self, ticker, date):

		return self.extract_sheet_by_date_group(date, ticker + "Greeks_Calls", 1)

	def extract_greeks_puts_date(self, ticker, date):

		return self.extract_sheet_by_date_group(date, ticker + "Greeks_Puts", 1)

	def extract_calls_date(self, ticker, date):

		return self.extract_sheet_by_date_group(date, ticker + "Calls", "K")	

	def extract_puts_date(self, ticker, date):

		return self.extract_sheet_by_date_group(date, ticker + "Puts", "K")	


	def extract_most_recent_quotes(self):

		n = self.xlsx["Prices"].max_row
		m = self.xlsx["Prices"].max_column

		columns = []

		for i in range(m):
			columns.append(i+1)

		return self.extract_xlsx_sheet_over_range("Prices", columns, [n])



	def get_estimated_greeks_by_option(self, ticker, T, option):


		now = datetime.datetime.now()
		date = now + datetime.timedelta(days = T)

		if option == "Call": 
			calls = self.extract_greeks_calls_date(ticker, date).drop(columns = "Root")

		else:
			calls = self.extract_greeks_puts_date(ticker, date)


		calls_date = self.parse_if_needed(calls.iloc[0,8])

		calls_last_adj_index = self.find_date_amongst_ordered(calls_date, ticker + "Historical_data", 7)
		calls_last_adj_close = self.xlsx[ticker + "Historical_data"][get_column_letter(6) + str(calls_last_adj_index)].value

		delta_T_calls = date.date() - self.parse_if_needed(self.xlsx[ticker + "Historical_data"][get_column_letter(7) + str(calls_last_adj_index)].value).date()
		delta_T_calls = delta_T_calls.days

		r = self.risk_free_rate
		r = (1 + r)**(delta_T_calls/365)

		strike = calls_last_adj_close * r

		frame = pd.DataFrame(calls.iloc[0,:])

		if option == "Call":

			frame.columns = [ticker]

		else:

			frame.columns = [ticker]

		x = calls["Strike"]

		for i in range(6):

			y = calls.iloc[:,i+1]

			spline = interp1d(x, y, kind='cubic')

			frame.iloc[i+1,0] = float(spline(strike))

		frame.iloc[7,0] = strike

		return frame



	def get_estimated_greeks(self, ticker, T):

		frame = self.get_estimated_greeks_by_option(ticker, T, "Call")
		frame2 = self.get_estimated_greeks_by_option(ticker, T, "Put")

		return {"Call" : frame , "Put" : frame2}


	def extract_estimated_greeks(self):

		Greeks = []

		for i in tqdm(self.stocks):

			try:

				G = self.get_estimated_greeks(i, T)

			except:

				continue

			else:

				Greeks.append(G)

		if len(Greeks) > 0:

			Greeks_Call = Greeks[0]["Call"]
			Greeks_Put = Greeks[0]["Put"]

			if len(Greeks) > 1:

				for i in range(len(Greeks)-1):

					Greeks_Call[Greeks[i]["Call"].columns.values[0]] = Greeks[i]["Call"]
					Greeks_Put[Greeks[i]["Put"].columns.values[0]] = Greeks[i]["Put"]


			return {"Call" : Greeks_Call , "Put" : Greeks_Put}


		else:

			print("ERROR: no data avaible.")
			return 0




	#-------------------------------

	#.xslx Data extractors

	#-------------------------------



	#-------------------------------

	#Unlocking and buying stocks

	#-------------------------------
	


	def unlock_stock(self, tickers):

		tickers = np.array(tickers)

		#Handles to case where self.stocks is empty

		if len(self.stocks) == 0:

			for i in tickers:

				try:
					data.get_quote_yahoo(i)

				except:
					print(i + " couldn't be found... Deleting")

				else:
					self.stocks = np.append(self.stocks, i)
					self.equity = np.append(self.equity, 0.0)
					self.shares = np.append(self.shares, 0.0)	

					n = self.xlsx["Stocks"].max_row + 1
					self.xlsx["Stocks"]["A" + str(n)] = i
					self.xlsx["Stocks"]["B" + str(n)] = 0

					self.xlsx["Prices"][get_column_letter(1) + str(1)] = i	

					self.xlsx.create_sheet(i)
					self.xlsx.create_sheet(i + "Calls")
					self.xlsx.create_sheet(i + "Puts")
					self.xlsx.create_sheet(i + "Historical_data")
					self.xlsx.create_sheet(i + "Greeks_Puts")
					self.xlsx.create_sheet(i + "Greeks_Calls")

					break	

		#Regular filling loop for non-empty self.stock

		for i in tickers:

			if(len(np.where(self.stocks == i)[0]) == 0):

				try:
					data.get_quote_yahoo(i)

				except:
					print(i + " couldn't be found... Deleting")

				else:
					self.stocks = np.append(self.stocks, i)
					self.equity = np.append(self.equity, 0.0)
					self.shares = np.append(self.shares, 0.0)	

					n = self.xlsx["Stocks"].max_row + 1
					self.xlsx["Stocks"]["A" + str(n)] = i
					self.xlsx["Stocks"]["B" + str(n)] = 0

					self.xlsx["Prices"][get_column_letter(self.xlsx["Prices"].max_column + 1) + str(1)] = i	


					self.xlsx.create_sheet(i)
					self.xlsx.create_sheet(i + "Calls")
					self.xlsx.create_sheet(i + "Puts")
					self.xlsx.create_sheet(i + "Historical_data")
					self.xlsx.create_sheet(i + "Greeks_Puts")
					self.xlsx.create_sheet(i + "Greeks_Calls")

			else:
				print(i + " is already in the portfolio")

		self.update()
		self.xlsx.save(self.file_location)




	def buy_single_shares(self, ticker, amount):

		index = np.where(self.stocks == ticker)[0]

		if len(index) != 0:
			self.shares[index] += amount

			self.xlsx["Stocks"]["B" + str(index[0] + 2)].value += amount		

			price = data.get_quote_yahoo(ticker)["price"][0]

			self.xlsx["Accounts"]["B1"].value -= price*amount

			self.assets["Debt"] -= price*amount

			self.equity[index] = price*self.shares[index]
			self.assets["Equity"] = np.sum(self.equity)
			self.xlsx["Accounts"]["A1"].value = self.assets["Equity"][0]

			print(str(amount) + " " + ticker + " shares were added.")
			print("")

			print("Current assets:")
			print("")

			print(self.assets)

			self.xlsx.save(self.file_location)


		else:
			print(ticker + " not found within the portfolio")




	def buy_vector_shares(self, amounts):


		if len(amounts) == len(self.shares):

			amounts = np.array(amounts)

			for index in tqdm(range(len(amounts))):

				try:

					self.xlsx["Stocks"]["B" + str(index + 2)].value += amounts[index]
					self.shares[index] += amounts[index]
				
					price = data.get_quote_yahoo(self.stocks[index])["price"][0]

					self.xlsx["Accounts"]["B1"].value -= price*amounts[index]
					self.assets["Debt"] -= price*amounts[index]

					self.equity[index] = price*self.shares[index]

				except:

					print(self.stocks[index] + " couldn't be updated; canceling transaction...")


			self.xlsx["Accounts"]["A1"].value = np.sum(self.equity)
			self.assets["Equity"] = self.xlsx["Accounts"]["A1"].value
			self.xlsx.save(self.file_location)

			print("Added new shares...")
			print("")

			print("Current assets:")
			print("")

			print(self.assets)

		else:
			print("Couldn't update shares: incorrect numbers of dimentions")
			print("Provided vector has " + str(len(amounts)) + " elements; Portfolio object has " + str(len(self.shares)))


	#-------------------------------

	#Unlocking and buying stocks

	#-------------------------------





	#-------------------------------

	#Updating and option purchasing

	#-------------------------------



	def update(self):

		#Obtain and store real time prices

		row = self.xlsx["Prices"].max_row + 1
		col = 0

		for i in tqdm(self.stocks):

			col += 1

			try:
				new_price = data.get_quote_yahoo(i)

			except:
				print("Couldn't update " + i + "...")

			else:
				self.xlsx["Prices"][get_column_letter(col) + str(row)] = new_price["price"][0]
				self.equity[col-1] = self.shares[col-1] * new_price["price"][0]
		 
		#Saves the time at which the request was made
		#These are the dates used to compute continuous interest on the Debt account

		now = datetime.datetime.now()
		date = str(now.year) + "-" + str(now.month) + "-" + str(now.day)

		n = self.xlsx["Dates"].max_row + 1

		self.xlsx["Dates"]["A" + str(n)].value = date
		self.xlsx["Dates"]["B" + str(n)].value = now.hour
		self.xlsx["Dates"]["C" + str(n)].value = now.minute
		self.xlsx["Dates"]["D" + str(n)].value = now.second

		self.assets["Equity"].value = np.sum(self.equity)

		try:

			old_date = self.xlsx["Dates"]["A" + str(n-1)].value + " " + str(self.xlsx["Dates"]["B" + str(n-1)].value) + ":" + str(self.xlsx["Dates"]["C" + str(n-1)].value) + ":" + str(self.xlsx["Dates"]["D" + str(n-1)].value)
			old_date = datetime.datetime.strptime(old_date, "%Y-%m-%d %H:%M:%S")

		except:
			print("Couldn't update debt...")

		else:
			r = np.log(1 + self.borrowing_rate)
			t = (now - old_date).total_seconds() / (60*60*24*365)

			self.xlsx["Accounts"]["B1"].value *= np.exp(t*r)
			self.assets["Debt"] *= np.exp(t*r)

		self.assets["Net assets"] = self.assets["Equity"] + self.assets["Debt"]

		n = self.xlsx["Dates"].max_row
		for k in np.arange(5,8):
			self.xlsx["Dates"][get_column_letter(k) + str(n)].value = self.assets.iloc[0,k-5]


		self.xlsx.save(self.file_location)

		print("Prices and Dates vectors successfully updated.")
		print("")

		print("Current assets:")
		print("")

		print(self.assets)


	def buy_call_put_from_data(self, ticker, expiration, strike, n_purchase, option_type):

		try:

			if option_type == "Call" or option_type == "call":

				frame = self.extract_calls_date(ticker, expiration)

			if option_type == "Put" or option_type == "put":

				frame = self.extract_puts_date(ticker, expiration)

		except:

			print("ERROR: something went wrong.")

			return 0

		else:

				strike_p = frame["Strike"]
				ask_p = frame["Last Price"]

				s_min = np.min(strike_p)
				s_max = np.max(strike_p)

				if strike >= s_min and strike <= s_max:

					try:

						spline = interp1d(strike_p, ask_p, kind='cubic')

						estimated_price = float(spline(strike))

						print("")
						print("Option price was estimated at: " + str(estimated_price) + "$")

						opt = [strike, n_purchase, expiration, option_type]
						column = ["Strike", "n_purchase", "Exercise date", "Type"]

						n = self.xlsx[ticker].max_row + 1

						#Save option and update balance

						for j in range(4):
							self.xlsx[ticker][get_column_letter(j+1) + str(1)].value = column[j]
							self.xlsx[ticker][get_column_letter(j+1) + str(n)].value = opt[j]

						self.assets["Debt"] -= estimated_price * n_purchase
						self.assets["Net assets"] -= estimated_price * n_purchase
						self.xlsx["Accounts"]["B1"].value -= estimated_price * n_purchase

						self.xlsx.save(self.file_location)

						print("")
						print("")
						print("")
						print("Purchase completed.")
						print("")
						print("Current Assets:")
						print("")
						print(self.assets)

					except:

						print("ERROR: something went wrong.")
						return 1



				else:

					print("Nonsensical strike price.")
					return -1

			



	def exercise_put_call(self):

		#Checks if some options should have been exercised based on the current date

		now = datetime.datetime.now()
		to_check = []
		to_rmv = []

		for i in self.stocks:

			n = self.xlsx[i].max_row

			if(n > 1):

				for j in range(n-1):

					then = self.xlsx[i]["C" + str(j+2)].value

					if type(then) == str:

						then = datetime.datetime.strptime(then, "%Y-%m-%d")

					if (then - now).total_seconds() < 0:

						opt = []

						for k in range(4):
							opt.append(self.xlsx[i][get_column_letter(k+1) + str(j+2)].value)

						opt.append(i)

						to_check.append(opt)
						to_rmv.append([i, j+2])


		#Exercise options

		if len(to_check) == 0:
			print("No call or put options to exercise")

		else:

			for i in to_check:

				to = i[2].strftime("%Y-%m-%d")
				info = data.DataReader(i[4], 'yahoo', to, to)

				if i[3] == "Call":

					if i[0] < info["Low"][0]:

						print("Buying " + str(i[1]) + " " + i[4] + " shares at " + str(i[0]) + " each.")

						index = np.where(self.stocks == i[4])[0][0]

						self.assets["Debt"] -= i[0] * i[1]
						self.xlsx["Accounts"]["B1"].value -= i[0] * i[1]

						self.assets["Net assets"] -= i[0] * i[1]
						
						self.shares[index] += i[1]
						self.xlsx["Stocks"]["B" + str(index+2)].value += i[1]


				else:

					if i[0] > info["High"][0]:

						print("Selling " + str(i[1]) + " " + i[4] + " shares at " + str(i[0]) + " each.")

						index = np.where(self.stocks == i[4])[0][0]

						self.assets["Debt"] += i[0] * i[1]
						self.xlsx["Accounts"]["B1"].value += i[0] * i[1]

						self.assets["Net assets"] += i[0] * i[1]

						self.shares[index] -= i[1]
						self.xlsx["Stocks"]["B" + str(index+2)].value -= i[1]

						#Handling the case where we didn't have enough shares to sell @ strike price

						if(self.shares[index] < 0):
							self.assets["Debt"] -= self.shares[index] * info["High"]
							self.xlsx["Accounts"]["B1"].value -= self.shares[index] * info["High"]
							self.shares[index] = 0
							self.xlsx["Stocks"]["B" + str(index+2)].value = 0



			#updating portfolio and removing used options

			self.update()


			to_rmv = np.array([j for i in to_rmv for j in i])

			for i in self.stocks:

				to_del = np.where(to_rmv == i)[0]

				if len(to_del) != 0:

					to_del += 1
					row_to_del = to_rmv[to_del].astype(int)
					row_to_del[::-1].sort()

					for j in row_to_del:
						self.xlsx[i].delete_rows(j)


			self.xlsx.save(self.file_location)





	#-------------------------------

	#Updating and option purchasing

	#-------------------------------





	#-------------------------------

	#Statistics

	#-------------------------------


	def get_historical_returns(self):

		frame = self.extract_adjusted_close()

		out = np.zeros(shape = (np.shape(frame)[0] - 1, np.shape(frame)[1]))

		for i in range(np.shape(frame)[1]):

			out[:,i] = (np.diff(frame.iloc[:,i]) / frame.iloc[1:,i]).iloc[:]

		return pd.DataFrame(out, columns = frame.columns, index = frame.index.values[:-1])



	def get_historical_log_returns(self):

		frame = self.extract_adjusted_close()

		frame_np = np.array(frame)

		out = np.zeros(shape = (np.shape(frame)[0] - 1, np.shape(frame)[1]))


		for i in range(np.shape(frame)[1]):

			out[:,i] = 	np.log(frame_np[1:,i] / frame_np[:-1,i])

		return pd.DataFrame(out, columns = frame.columns, index = frame.index.values[:-1])


	def get_historical_returns_over_range(self, frm, to):

		frame = self.extract_adjusted_close_over_range(frm, to)

		out = np.zeros(shape = (np.shape(frame)[0] - 1, np.shape(frame)[1]))

		for i in range(np.shape(frame)[1]):

			out[:,i] = (np.diff(frame.iloc[:,i]) / frame.iloc[1:,i]).iloc[:]

		return pd.DataFrame(out, columns = frame.columns, index = frame.index.values[:-1])



	def get_historical_log_returns_over_range(self, frm, to):

		frame = self.extract_adjusted_close_over_range(frm, to)

		frame_np = np.array(frame)

		out = np.zeros(shape = (np.shape(frame)[0] - 1, np.shape(frame)[1]))


		for i in range(np.shape(frame)[1]):

			out[:,i] = 	np.log(frame_np[1:,i] / frame_np[:-1,i])

		return pd.DataFrame(out, columns = frame.columns, index = frame.index.values[:-1])



	def get_covariance_matrix_from_close_over_range(self, frm, to):

		return self.extract_adjusted_close_over_range(frm, to).cov()

	def get_correlation_matrix_from_close_over_range(self, frm, to):

		return self.extract_adjusted_close_over_range(frm, to).corr()


	def get_covariance_matrix_from_returns_over_range(self, frm, to):

		return self.get_historical_returns_over_range(frm, to).cov()

	def get_correlation_matrix_from_returns_over_range(self, frm, to):

		return self.get_historical_returns_over_range(frm, to).corr()


	def get_covariance_matrix_from_log_returns_over_range(self, frm, to):

		return self.get_historical_log_returns_over_range(frm, to).cov()

	def get_correlation_matrix_from_log_returns_over_range(self, frm, to):

		return self.get_historical_log_returns_over_range(frm, to).corr()


	def get_covariance_matrix_from_close(self):

		return self.extract_adjusted_close().cov()

	def get_correlation_matrix_from_close(self):

		return self.extract_adjusted_close().corr()


	def get_covariance_matrix_from_returns(self):

		return self.get_historical_returns().cov()

	def get_correlation_matrix_from_returns(self):

		return self.get_historical_returns().corr()


	def get_covariance_matrix_from_log_returns(self):

		return self.get_historical_log_returns().cov()

	def get_correlation_matrix_from_log_returns(self):

		return self.get_historical_log_returns().corr()

	#-------------------------------

	#Statistics

	#-------------------------------




	#-------------------------------

	#Weights

	#-------------------------------


	def get_lowest_variance_pf_weights(self):

		v_cov = self.get_covariance_matrix_from_log_returns()

		try:

			v_cov_inv = np.linalg.inv(v_cov)

		except:

			v_cov_inv = np.linalg.pinv(v_cov)

		J = np.ones((np.shape(v_cov)[1], 1))
		v = np.matmul(v_cov_inv, J)
		v /= np.dot(np.transpose(J),v)

		return np.transpose(v)

	

	def get_lowest_variance_pf_over_range(self, frm, to):

		v_cov = self.get_covariance_matrix_from_log_returns_over_range(frm, to)
		v_cov_inv = np.linalg.inv(v_cov)

		J = np.ones((np.shape(v_cov)[1], 1))
		v = np.matmul(v_cov_inv, J)
		v /= np.dot(np.transpose(J),v)

		return np.transpose(v)



	
	def get_eigen_pf_weights(self):

		v_cov = self.get_covariance_matrix_from_log_returns()

		l, v = np.linalg.eig(v_cov)

		return v[0] / sum(np.abs(v[0]))



	def get_eigen_pf_weights_over_range(self, frm, to):

		v_cov = self.get_covariance_matrix_from_log_returns_over_range(frm, to)

		l, v = np.linalg.eig(v_cov)

		return v[0] / sum(np.abs(v[0]))




	def get_sharpe_weights(self):

		#Extract needed objects

		data = self.get_historical_log_returns()

		v_cov = data.cov()

		returns = np.zeros(np.shape(data)[1])

		for i in range(len(returns)):

			returns[i] = np.exp(np.sum(data.iloc[:,i])) - 1

		w = np.zeros(len(returns))

		for i in range(len(w)):
			w[i] = 1

		w /= len(w)


		#Define functions needed to perform multivariate Newton-Rhapson method

		def obj_function(w, returns, v_cov):

			return (np.dot(returns, w) / np.sqrt(np.dot(np.transpose(w), np.matmul(v_cov, w))))


		def get_gradient(w, returns, v_cov):

	
			return returns - (np.dot(returns, w) / np.dot(np.transpose(w), np.matmul(v_cov, w))) * np.matmul(v_cov, w)


		def get_hessian(w, returns, v_cov):

			a = np.dot(returns, w)
			b = np.sqrt(np.dot(np.transpose(w), np.matmul(v_cov, w)))
			c = np.matmul(np.transpose(w), v_cov)


			(c / b**2) * 2 * v_cov

			return np.outer(-(returns / b - (a / b**2) * 2 * w) , c) - (a /b) * v_cov


		def get_direction(w, returns, v_cov):

			H = get_hessian(w, returns, v_cov)

			try:

				H = np.linalg.inv(H)

			except:

				H = np.linalg.pinv(H)

			return  np.matmul(H, get_gradient(w, returns, v_cov))


		#Optimise

		old = obj_function(w, returns, v_cov)
		new = 2*old

		k = 0

		while True:

			k += 1

			old = new
			w -= get_direction(w, returns, v_cov)
			new = obj_function(w, returns, v_cov)

			if np.abs(new / old - 1) < 0.00000001:

				break

		print("Converged after " + str(k) + " iterations.")

		return w / sum(np.abs(w))




	def get_sharpe_weights_over_range(self, frm, to):

		#Extract needed objects

		data = self.get_historical_log_returns_over_range(frm, to)

		v_cov = data.cov()

		returns = np.zeros(np.shape(data)[1])

		for i in range(len(returns)):

			returns[i] = np.exp(np.sum(data.iloc[:,i])) - 1

		w = np.zeros(len(returns))

		for i in range(len(w)):
			w[i] = 1

		w /= len(w)


		#Define functions needed to perform multivariate Newton-Rhapson method

		def obj_function(w, returns, v_cov):

			return (np.dot(returns, w) / np.sqrt(np.dot(np.transpose(w), np.matmul(v_cov, w))))


		def get_gradient(w, returns, v_cov):

	
			return returns - (np.dot(returns, w) / np.dot(np.transpose(w), np.matmul(v_cov, w))) * np.matmul(v_cov, w)


		def get_hessian(w, returns, v_cov):

			a = np.dot(returns, w)
			b = np.sqrt(np.dot(np.transpose(w), np.matmul(v_cov, w)))
			c = np.matmul(np.transpose(w), v_cov)


			(c / b**2) * 2 * v_cov

			return np.outer(-(returns / b - (a / b**2) * 2 * w) , c) - (a /b) * v_cov


		def get_direction(w, returns, v_cov):

			H = get_hessian(w, returns, v_cov)

			try:

				H = np.linalg.inv(H)

			except:

				H = np.linalg.pinv(H)

			return  np.matmul(H, get_gradient(w, returns, v_cov))


		#Optimise

		old = obj_function(w, returns, v_cov)
		new = 2*old

		k = 0

		while True:

			k += 1

			old = new
			w -= get_direction(w, returns, v_cov)
			new = obj_function(w, returns, v_cov)

			if np.abs(new / old - 1) < 0.00000001:

				break

		print("Converged after " + str(k) + " iterations.")

		return w / sum(np.abs(w))






	def get_scaled_historical_close(self):

		prices = self.extract_adjusted_close()

		for j in range(np.shape(prices)[1]):

			prices.iloc[:,j] /= prices.iloc[0,j]

		return prices


	def get_scaled_historical_close_over_range(self, frm, to):

		prices = self.extract_adjusted_close_over_range(frm, to)

		for j in range(np.shape(prices)[1]):

			prices.iloc[:,j] /= prices.iloc[0,j]

		return prices




	def get_pairwise_Mahalanobis_distances(self, matrix, axis):
		
		if axis == 1:

			matrix = np.transpose(matrix)

		v_cov = matrix.cov()

		L = np.linalg.cholesky(v_cov)
		L = np.linalg.inv(L)

		n = np.shape(matrix)[0]

		out = np.zeros(shape = (n, n))

		print("")
		print("Computing pairwise distances...")

		for i in tnrange(n, desc = "Outer loop"):
			for j in range(i+1):

				x = matrix.iloc[i,:]
				y = matrix.iloc[j,:]

				out[i,j] = np.linalg.norm(np.matmul(L,(x-y)))
				out[j,i] = out[i,j]


		out = pd.DataFrame(out, columns = matrix.index, index = matrix.index)
		

		return out



		def get_Mahalanobis_distances(self, matrix, axis):

			if axis == 1:

				matrix = np.transpose(matrix)

			v_cov = matrix.cov()

			for j in range(np.shape(matrix)[1]):

				matrix.iloc[:,j] -= matrix.iloc[:,j].mean()

			L = np.linalg.cholesky(v_cov)
			L = np.linalg.inv(L)

			norms = np.matmul(L, np.transpose(matrix))

			n = np.shape(matrix)[0]

			out = np.zeros(shape = (n, 1))

			print("")
			print("Computing row-wise distances from mean...")

			for j in tqdm(range(n)):
				out[j] = np.linalg.norm(norms[:,j])

			out = pd.DataFrame(out, index = matrix.index, columns = ["Malh Distances"])

			return out









	def get_ANNUAL_implied_return_and_volatility(self, ticker, frm, to):


		frm_index = self.find_date_amongst_ordered(frm, ticker + "Historical_data", 7)
		to_index = self.find_date_amongst_ordered(to, ticker + "Historical_data", 7)
		
		rows = []
		for i in range(to_index - frm_index + 1):

			rows.append(frm_index + i)

		columns = [6]

		frame = self.extract_xlsx_sheet_over_range(ticker + "Historical_data", columns, rows)
		dates = self.extract_xlsx_sheet_over_range(ticker + "Historical_data", [7], rows)

		ret = np.zeros(np.shape((frame))[0] - 1)

		for i in range(np.shape(ret)[0]):
			ret[i] = frame.iloc[i+1,0] / frame.iloc[i,0]

		ret = np.log(ret)

		for i in range(len(ret)):

			delta = self.date_diff_seconds(dates.iloc[i+1,0], dates.iloc[i,0]) / 86400 
			ret[i] /= delta

		delta = 1 / 365

		mean = np.mean(ret)

		vol = np.sum((ret - mean)**2) / (delta*(len(ret)-1))

		ret = mean / delta + vol**2 / 2

		out = pd.DataFrame(data = [[ret, vol]], columns = ["Return", "Volatility"])

		#print("")
		#print(ticker)
		#print("")
		#print("from " + frm + " to " + to)
		#print("")
		#print(out)


		return out





	def analyse_time_frame(self, frm, T, to):

		frm = self.parse_if_needed(frm)
		to = self.parse_if_needed(to)

		print("")
		print("Extracting Adj Close...")
		print("")

		adj_close = self.extract_adjusted_close_over_range(frm, to)

		unit_close = np.zeros(shape=np.shape(adj_close))

		for j in range(np.shape(unit_close)[1]):

			unit_close[:,j] = adj_close.iloc[:,j] / adj_close.iloc[0,j]

		unit_close = pd.DataFrame(unit_close, columns = adj_close.columns, index = adj_close.index)

		print(unit_close.plot())

		print("Computing weights...")
		print("")

		min_var_w = self.get_lowest_variance_pf_over_range(frm, to)
		eigen_w = self.get_eigen_pf_weights_over_range(frm, to)
		sr_w = self.get_sharpe_weights_over_range(frm, to)
		pf_w = adj_close.iloc[0,:] / np.sum(adj_close.iloc[0,:])

		w_frame = np.zeros(shape = (3, len(sr_w)))
		w_frame[0,:] = min_var_w 
		w_frame[1,:] = eigen_w
		w_frame[2,:] = sr_w

		w_frame = pd.DataFrame(w_frame, columns = unit_close.columns, index = ["Min Var", "PC", "Sharpe"])

		print("Computing weighted returns...")
		print("")


		pf_by_w = np.zeros(shape = (np.shape(unit_close)[0], 4))

		for i in tqdm(range(np.shape(unit_close)[0])):

			pf_by_w[i,0] = np.dot(min_var_w, unit_close.iloc[i,:])
			pf_by_w[i,1] = np.dot(eigen_w, unit_close.iloc[i,:])
			pf_by_w[i,2] = np.dot(sr_w, unit_close.iloc[i,:])
			pf_by_w[i,3] = np.dot(pf_w, unit_close.iloc[i,:])

		pf_by_w = pd.DataFrame(pf_by_w, columns = ["Min Var", "PC", "Sharpe", "Current"], index = adj_close.index)

		print(pf_by_w.plot())


		print("Computing statistics...")
		print("")

		cor_mat = self.get_correlation_matrix_from_returns_over_range(frm, to)
		v_cov = self.get_covariance_matrix_from_log_returns_over_range(frm, to)

		print(sns.heatmap(cor_mat, vmax=1, vmin = -1, square=True))

		log_ret = self.get_historical_log_returns_over_range(frm, to)

		Malh_dists = self.get_Mahalanobis_distances(log_ret, ax = 0)
		print(Malh_dists.plot())


		print("Extracting Greeks...")
		print("")

		Greeks = self.extract_estimated_greeks()

		vol_ret_frame = extract_all_ret_vol_over_range(frm, to)


		return {"Adj_Close" : adj_close, "Unit_Close" : unit_close, "Weights" : w_frame, 
			"Weighted_Pf" : pf_by_w, "Cor" : cor_mat, "V_Cov" : v_cov, "Log_ret" : log_ret, "Malh_Dist" : Malh_dists, "Greeks" : Greeks, "Vol_Ret" : vol_ret_frame}



	def extract_all_ret_vol_over_range(self, frm, to):

		vols = []
		rets = []
		cols = []

		for i in self.stocks:

			try:

				v = self.get_ANNUAL_implied_return_and_volatility(i, frm, to)

			except:
				continue

			else:

				rets.append(v.iloc[0,0])
				vols.append(v.iloc[0,1])
				cols.append(i)


		if len(vols) > 0:

			vol_ret_frame = pd.DataFrame([rets, vols], columns = cols, index = ["Return", "Volatility"])
			return vol_ret_frame


		else:
			print("ERROR: no data, returning 0 instead.")
			return 0

















































		








	





































